import cgi
import io
import os
import sys
from http.server import BaseHTTPRequestHandler
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# Path to the template stored in the repo
TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "backend" / "templates" / "Consolidated Mentor-Mentee form.xlsx"

# -----------------------------------------------------------------
# CELL MAP — maps each form field name to its Excel cell address.
# These are estimated positions based on the template layout.
# Adjust after visually verifying the template in Excel/LibreOffice.
# -----------------------------------------------------------------
CELL_MAP = {
    # Section 1 – Student Profile
    "student_name":         "E10",
    "roll_number":          "E11",
    "year_of_admission":    "E12",
    "branch":               "E13",
    "date_of_birth":        "E14",
    "student_aadhaar":      "E15",
    "student_appar":        "E16",
    # Section 2 – Parents
    "father_name":          "E18",
    "mother_name":          "E19",
    # Section 3 – Communication Address
    "comm_address_line1":   "E21",
    "comm_address_line2":   "E22",
    "comm_address_line3":   "E23",
    "comm_address_line4":   "E24",
    # Section 4 – Permanent Address
    "perm_address_line1":   "E26",
    "perm_address_line2":   "E27",
    "perm_address_line3":   "E28",
    "perm_address_line4":   "E29",
    # Section 5 – Contact
    "father_mobile":        "E31",
    "father_email":         "K31",
    "mother_mobile":        "E32",
    "mother_email":         "K32",
    "student_mobile":       "E33",
    "student_email":        "K33",
    # Section 6 – Academic Background
    "school_upto_x":        "H36",
    "intermediate_college": "H37",
    "year_of_completion":   "H38",
    "percentage_inter":     "H39",
    "eamcet_rank":          "H40",
    "category_admission":   "H41",
    # Section 7 – Residing
    "residing_type":        "E43",
    "hostel_address":       "E46",
    # Section 8 – Identification
    "id_mark_1":            "E48",
    "id_mark_2":            "E49",
    # Section 9 – Hobbies
    "hobbies":              "E51",
    # Section 10 – Health
    "health_issues":        "E53",
    # Section 11 – Transport
    "transport_mode":       "E55",
    # Section 12 – Parent Background
    "father_occupation":    "H57",
    "father_edu_qual":      "H58",
    "father_income":        "H59",
    "mother_occupation":    "H61",
    "mother_edu_qual":      "H62",
    "mother_income":        "H63",
    # Section 13 – Career
    "computer_courses":     "H66",
    "competitive_exams":    "H67",
}


def parse_multipart(environ):
    """Parse multipart/form-data from WSGI-style environ."""
    content_type = environ.get("CONTENT_TYPE", "")
    content_length = int(environ.get("CONTENT_LENGTH", 0) or 0)
    body = environ["wsgi.input"].read(content_length)
    fp = io.BytesIO(body)
    environ2 = {"REQUEST_METHOD": "POST", "CONTENT_TYPE": content_type, "CONTENT_LENGTH": str(content_length)}
    # Use cgi.FieldStorage for parsing
    fs = cgi.FieldStorage(fp=fp, environ=environ2, keep_blank_values=True)
    return fs


class handler(BaseHTTPRequestHandler):

    def do_OPTIONS(self):
        self.send_response(200)
        self._cors()
        self.end_headers()

    def do_POST(self):
        try:
            content_type = self.headers.get("Content-Type", "")
            content_length = int(self.headers.get("Content-Length", 0))
            body = self.rfile.read(content_length)

            # Parse multipart form
            fp = io.BytesIO(body)
            environ = {
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": content_type,
                "CONTENT_LENGTH": str(content_length),
                "wsgi.input": fp,
            }
            fs = cgi.FieldStorage(
                fp=io.BytesIO(body),
                headers=self.headers,
                environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": content_type},
                keep_blank_values=True,
            )

            def get(name):
                item = fs.getvalue(name)
                if isinstance(item, bytes):
                    return item.decode("utf-8", errors="ignore")
                return item or ""

            # Load template
            if not TEMPLATE_PATH.exists():
                self._error(500, f"Template not found: {TEMPLATE_PATH}")
                return

            wb = load_workbook(TEMPLATE_PATH)
            ws = wb[wb.sheetnames[0]]

            # Write all text fields
            for field, cell in CELL_MAP.items():
                val = get(field)
                if val:
                    ws[cell] = val

            # Embed photo if provided
            if "photo" in fs and fs["photo"].filename:
                photo_bytes = fs["photo"].file.read()
                if photo_bytes:
                    try:
                        img = XLImage(io.BytesIO(photo_bytes))
                        img.width = 100
                        img.height = 120
                        ws.add_image(img, "L9")  # top-right area of student profile
                    except Exception:
                        pass  # photo embed is non-critical

            # Save to in-memory buffer
            out_buf = io.BytesIO()
            wb.save(out_buf)
            out_buf.seek(0)
            excel_bytes = out_buf.read()

            roll = get("roll_number") or "student"
            name = get("student_name").replace(" ", "_") or "form"
            filename = f"{roll}_{name}.xlsx"

            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
            self.send_header("Content-Length", str(len(excel_bytes)))
            self._cors()
            self.end_headers()
            self.wfile.write(excel_bytes)

        except Exception as e:
            self._error(500, str(e))

    def _cors(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "*")

    def _error(self, code, msg):
        body = msg.encode()
        self.send_response(code)
        self.send_header("Content-Type", "text/plain")
        self.send_header("Content-Length", str(len(body)))
        self._cors()
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, format, *args):
        pass  # suppress default logging
