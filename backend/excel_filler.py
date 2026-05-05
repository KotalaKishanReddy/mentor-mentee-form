from pathlib import Path
from typing import Dict, Any
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = BASE_DIR / "templates" / "Consolidated Mentor-Mentee form.xlsx"
OUTPUT_DIR = BASE_DIR / "generated"

CELL_MAP: Dict[str, str] = {
    "student_name": "E10",
}


def fill_template(data: Dict[str, Any], output_filename: str) -> Path:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template not found at {TEMPLATE_PATH}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[wb.sheetnames[0]]

    for key, value in data.items():
        cell_ref = CELL_MAP.get(key)
        if cell_ref and value not in (None, ""):
            ws[cell_ref] = value

    output_path = OUTPUT_DIR / output_filename
    wb.save(output_path)
    return output_path
